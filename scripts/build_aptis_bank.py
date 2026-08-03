#!/usr/bin/env python3
"""Build and validate the Aptis Core + Reading Learning Graph snapshot."""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

REQUIRED_STATUS = "PUBLISHED_FINAL"
EXPECTED_GRAMMAR = 1000
EXPECTED_VOCABULARY = 1000
EXPECTED_READING = 290
EXPECTED_READING_TESTS = 10
EXPECTED_READING_TASKS = 50
EXPECTED_READING_UNITS = 300
EXPECTED_ITEMS = 2290


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--version", default="3.0.0")
    parser.add_argument("--source-sheet-id", default="")
    return parser.parse_args()


def snake_case(value: str) -> str:
    value = re.sub(r"[^A-Za-z0-9]+", "_", value.strip())
    return re.sub(r"_+", "_", value).strip("_").lower()


def read_rows(workbook: Any, sheet_name: str, key_field: str | None = None) -> list[dict[str, Any]]:
    if sheet_name not in workbook.sheetnames:
        raise ValueError(f"Missing worksheet: {sheet_name}")
    worksheet = workbook[sheet_name]
    iterator = worksheet.iter_rows(values_only=True)
    headers_raw = next(iterator)
    headers = [str(value) if value is not None else "" for value in headers_raw]
    if key_field and key_field not in headers:
        raise ValueError(f"{sheet_name}: missing key field {key_field}")
    key_index = headers.index(key_field) if key_field else 0
    output: list[dict[str, Any]] = []
    for row in iterator:
        if not row or row[key_index] in (None, ""):
            continue
        output.append({header: row[index] for index, header in enumerate(headers) if header})
    return output


def require_fields(row: dict[str, Any], fields: list[str], row_id: str) -> None:
    missing = [field for field in fields if row.get(field) in (None, "")]
    if missing:
        raise ValueError(f"{row_id}: missing fields: {', '.join(missing)}")


def require_unique_ids(rows: list[dict[str, Any]], field: str, label: str) -> None:
    values = [str(row[field]) for row in rows]
    if len(values) != len(set(values)):
        raise ValueError(f"{label}: duplicate values in {field}")


def build_grammar(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    required = [
        "ID", "Section", "Item Type", "Topic", "Level", "Question",
        "Option A", "Option B", "Option C", "Correct",
        "Explanation EN", "Explanation VI", "Difficulty", "Status",
        "Primary Concept ID", "Adaptive Bucket", "Graph Status",
    ]
    output = []
    for row in rows:
        row_id = str(row.get("ID"))
        require_fields(row, required, row_id)
        options = [str(row["Option A"]), str(row["Option B"]), str(row["Option C"])]
        correct = str(row["Correct"])
        if correct not in "ABC":
            raise ValueError(f"{row_id}: invalid answer label {correct}")
        if len(set(options)) != 3:
            raise ValueError(f"{row_id}: duplicate answer options")
        if str(row["Status"]) != REQUIRED_STATUS:
            raise ValueError(f"{row_id}: status must be {REQUIRED_STATUS}")
        output.append({
            "id": row_id,
            "section": str(row["Section"]),
            "item_type": str(row["Item Type"]),
            "topic": str(row["Topic"]),
            "level": str(row["Level"]),
            "question": str(row["Question"]),
            "options": options,
            "correct": correct,
            "explanation_en": str(row["Explanation EN"]),
            "explanation_vi": str(row["Explanation VI"]),
            "difficulty": str(row["Difficulty"]),
            "status": str(row["Status"]),
            "primary_concept_id": str(row["Primary Concept ID"]),
            "topic_node_id": str(row.get("Topic Node ID") or ""),
            "skill": str(row.get("Skill") or ""),
            "canonical_target": str(row.get("Canonical Target") or ""),
            "context_cue": str(row.get("Context Cue") or ""),
            "ambiguity_risk": str(row.get("Ambiguity Risk") or "LOW"),
            "prerequisite_ids": str(row.get("Prerequisite IDs") or ""),
            "related_concepts": str(row.get("Related Concepts") or ""),
            "adaptive_bucket": str(row["Adaptive Bucket"]),
            "mastery_weight": float(row.get("Mastery Weight") or 1.0),
            "review_priority": str(row.get("Review Priority") or "NORMAL"),
            "template_key": str(row.get("Template Key") or ""),
            "graph_status": str(row["Graph Status"]),
        })
    if len(output) != EXPECTED_GRAMMAR:
        raise ValueError(f"Expected {EXPECTED_GRAMMAR} grammar items, got {len(output)}")
    return output


def build_vocabulary(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    required = [
        "ID", "Section", "Item Type", "Subtype", "Level", "Prompt",
        "Option A", "Option B", "Option C", "Correct", "Correct Value",
        "Explanation EN", "Explanation VI", "Difficulty", "Status",
        "Primary Concept ID", "Adaptive Bucket", "Graph Status",
    ]
    output = []
    for row in rows:
        row_id = str(row.get("ID"))
        require_fields(row, required, row_id)
        item_type = str(row["Item Type"])
        if item_type == "bank_match" and row.get("Group ID") in (None, ""):
            raise ValueError(f"{row_id}: missing fields: Group ID")
        correct = str(row["Correct"])
        labels = "ABCDEFGHIJ" if item_type == "bank_match" else "ABC"
        option_count = 10 if item_type == "bank_match" else 3
        options = [str(row[f"Option {label}"]) for label in labels[:option_count]]
        if correct not in labels[:option_count]:
            raise ValueError(f"{row_id}: invalid answer label {correct}")
        if len(set(options)) != option_count:
            raise ValueError(f"{row_id}: duplicate answer options")
        correct_value = str(row["Correct Value"])
        if options[labels.index(correct)] != correct_value:
            raise ValueError(f"{row_id}: answer label does not match Correct Value")
        if str(row["Status"]) != REQUIRED_STATUS:
            raise ValueError(f"{row_id}: status must be {REQUIRED_STATUS}")
        item = {
            "id": row_id,
            "section": str(row["Section"]),
            "item_type": item_type,
            "subtype": str(row["Subtype"]),
            "group_id": "" if row.get("Group ID") in (None, "") else str(row["Group ID"]),
            "level": str(row["Level"]),
            "prompt": str(row["Prompt"]),
            "correct": correct,
            "correct_value": correct_value,
            "explanation_en": str(row["Explanation EN"]),
            "explanation_vi": str(row["Explanation VI"]),
            "difficulty": str(row["Difficulty"]),
            "status": str(row["Status"]),
            "primary_concept_id": str(row["Primary Concept ID"]),
            "secondary_concept_id": str(row.get("Secondary Concept ID") or ""),
            "skill": str(row.get("Skill") or ""),
            "family": str(row.get("Family") or ""),
            "canonical_target": str(row.get("Canonical Target") or correct_value),
            "pattern": str(row.get("Pattern") or ""),
            "context_cue": str(row.get("Context Cue") or ""),
            "ambiguity_risk": str(row.get("Ambiguity Risk") or "LOW"),
            "prerequisite_ids": str(row.get("Prerequisite IDs") or ""),
            "related_concepts": str(row.get("Related Concepts") or ""),
            "adaptive_bucket": str(row["Adaptive Bucket"]),
            "mastery_weight": float(row.get("Mastery Weight") or 1.0),
            "review_priority": str(row.get("Review Priority") or "NORMAL"),
            "template_key": str(row.get("Template Key") or ""),
            "graph_status": str(row["Graph Status"]),
        }
        item["bank_options" if item_type == "bank_match" else "options"] = options
        output.append(item)
    if len(output) != EXPECTED_VOCABULARY:
        raise ValueError(f"Expected {EXPECTED_VOCABULARY} vocabulary items, got {len(output)}")
    return output


def build_reading_items(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    required = [
        "ID", "Section", "Test ID", "Group ID", "Part", "Item Type", "Level",
        "Prompt", "Option A", "Correct", "Correct Value", "Explanation EN",
        "Explanation VI", "Difficulty", "Status", "Primary Concept ID",
        "Secondary Concept ID", "Skill", "Adaptive Bucket", "Graph Status",
    ]
    output: list[dict[str, Any]] = []
    for row in rows:
        row_id = str(row.get("ID"))
        require_fields(row, required, row_id)
        labels = "ABCDEFGH"
        present_labels = [label for label in labels if row.get(f"Option {label}") not in (None, "")]
        options = [str(row[f"Option {label}"]) for label in present_labels]
        correct = str(row["Correct"])
        if not 3 <= len(options) <= 8:
            raise ValueError(f"{row_id}: expected 3-8 reading options, got {len(options)}")
        if correct not in present_labels:
            raise ValueError(f"{row_id}: invalid answer label {correct}")
        if len(set(options)) != len(options):
            raise ValueError(f"{row_id}: duplicate answer options")
        if options[present_labels.index(correct)] != str(row["Correct Value"]):
            if int(row["Part"]) != 3:
                raise ValueError(f"{row_id}: answer label does not match Correct Value")
        if str(row["Status"]) != REQUIRED_STATUS:
            raise ValueError(f"{row_id}: status must be {REQUIRED_STATUS}")
        if str(row.get("Ambiguity Risk") or "LOW") != "LOW":
            raise ValueError(f"{row_id}: reading exam item must have LOW ambiguity risk")
        output.append({
            "id": row_id,
            "section": str(row["Section"]),
            "test_id": str(row["Test ID"]),
            "group_id": str(row["Group ID"]),
            "part": int(row["Part"]),
            "item_type": str(row["Item Type"]),
            "level": str(row["Level"]),
            "prompt": str(row["Prompt"]),
            "options": options,
            "correct": correct,
            "correct_value": str(row["Correct Value"]),
            "explanation_en": str(row["Explanation EN"]),
            "explanation_vi": str(row["Explanation VI"]),
            "difficulty": str(row["Difficulty"]),
            "status": str(row["Status"]),
            "primary_concept_id": str(row["Primary Concept ID"]),
            "secondary_concept_id": str(row["Secondary Concept ID"]),
            "skill": str(row["Skill"]),
            "family_subskill": str(row.get("Family/Subskill") or ""),
            "context_cue": str(row.get("Context Cue") or ""),
            "ambiguity_risk": str(row.get("Ambiguity Risk") or "LOW"),
            "adaptive_bucket": str(row["Adaptive Bucket"]),
            "mastery_weight": float(row.get("Mastery Weight") or 1.0),
            "review_priority": str(row.get("Review Priority") or "NORMAL"),
            "template_key": str(row.get("Template Key") or ""),
            "graph_status": str(row["Graph Status"]),
            "source_basis": str(row.get("Source Basis") or ""),
        })
    if len(output) != EXPECTED_READING:
        raise ValueError(f"Expected {EXPECTED_READING} reading items, got {len(output)}")
    require_unique_ids(output, "id", "READING_ITEMS")
    return output


def normalise_table(rows: list[dict[str, Any]]) -> list[dict[str, Any]]:
    return [{snake_case(key): value for key, value in row.items()} for row in rows]


def write_json(path: Path, value: Any) -> bytes:
    payload = json.dumps(value, ensure_ascii=False, separators=(",", ":")).encode("utf-8")
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_bytes(payload)
    temporary.replace(path)
    return payload


def main() -> int:
    args = parse_args()
    if not args.input.exists():
        raise FileNotFoundError(args.input)
    args.output_dir.mkdir(parents=True, exist_ok=True)

    workbook = load_workbook(args.input, read_only=True, data_only=True)
    try:
        grammar = build_grammar(read_rows(workbook, "GRAMMAR_1000", "ID"))
        vocabulary = build_vocabulary(read_rows(workbook, "VOCABULARY_1000", "ID"))
        reading_tests = normalise_table(read_rows(workbook, "READING_TESTS", "Test ID"))
        reading_tasks = normalise_table(read_rows(workbook, "READING_TASKS", "Group ID"))
        reading_units = normalise_table(read_rows(workbook, "READING_UNITS", "Unit ID"))
        reading_items = build_reading_items(read_rows(workbook, "READING_ITEMS", "ID"))
        concepts = normalise_table(read_rows(workbook, "CONCEPT_REGISTRY", "Concept ID"))
        edges = normalise_table(read_rows(workbook, "CONCEPT_EDGES", "Source Concept ID"))
        item_graph = normalise_table(read_rows(workbook, "ITEM_GRAPH", "Item ID"))
        adaptive_rules = normalise_table(read_rows(workbook, "ADAPTIVE_RULES", "Rule ID"))
        learner_state = normalise_table(read_rows(workbook, "LEARNER_STATE", "Learner"))
    finally:
        workbook.close()

    if len(reading_tests) != EXPECTED_READING_TESTS:
        raise ValueError(f"Expected {EXPECTED_READING_TESTS} reading tests, got {len(reading_tests)}")
    if len(reading_tasks) != EXPECTED_READING_TASKS:
        raise ValueError(f"Expected {EXPECTED_READING_TASKS} reading tasks, got {len(reading_tasks)}")
    if len(reading_units) != EXPECTED_READING_UNITS:
        raise ValueError(f"Expected {EXPECTED_READING_UNITS} reading units, got {len(reading_units)}")
    if len(item_graph) != EXPECTED_ITEMS:
        raise ValueError(f"Expected {EXPECTED_ITEMS} item graph mappings, got {len(item_graph)}")
    if len(concepts) < 880:
        raise ValueError(f"Concept registry is unexpectedly small: {len(concepts)}")
    if len(edges) < 1325:
        raise ValueError(f"Concept edge set is unexpectedly small: {len(edges)}")

    all_questions = grammar + vocabulary + reading_items
    question_ids = {item["id"] for item in all_questions}
    if len(question_ids) != EXPECTED_ITEMS:
        raise ValueError("Question IDs are not globally unique")
    graph_ids = {str(item["item_id"]) for item in item_graph}
    if question_ids != graph_ids:
        missing = sorted(question_ids - graph_ids)[:10]
        extra = sorted(graph_ids - question_ids)[:10]
        raise ValueError(f"ITEM_GRAPH mismatch; missing={missing}, extra={extra}")

    concept_ids = {str(item["concept_id"]) for item in concepts}
    unknown = sorted({str(item["primary_concept_id"]) for item in item_graph} - concept_ids)
    if unknown:
        raise ValueError(f"Unknown primary concepts in ITEM_GRAPH: {unknown[:10]}")

    generated = {
        "grammar.json": grammar,
        "vocabulary.json": vocabulary,
        "reading-tests.json": reading_tests,
        "reading-tasks.json": reading_tasks,
        "reading-units.json": reading_units,
        "reading-items.json": reading_items,
        "concepts.json": concepts,
        "edges.json": edges,
        "item_graph.json": item_graph,
        "adaptive_rules.json": adaptive_rules,
        "learner_state.json": learner_state,
    }
    payloads = [write_json(args.output_dir / name, data) for name, data in generated.items()]
    revision = hashlib.sha256(b"".join(payloads)).hexdigest()[:16]
    source_sha256 = hashlib.sha256(args.input.read_bytes()).hexdigest()

    generated_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    reading_manifest = {
        "version": args.version,
        "status": "READING_B2_PLUS",
        "question_status": REQUIRED_STATUS,
        "generated_at": generated_at,
        "source_sheet_id": args.source_sheet_id,
        "test_count": len(reading_tests),
        "task_count": len(reading_tasks),
        "unit_count": len(reading_units),
        "item_count": len(reading_items),
        "part_counts": {
            str(part): sum(1 for item in reading_items if item["part"] == part)
            for part in range(1, 5)
        },
    }
    write_json(args.output_dir / "reading-manifest.json", reading_manifest)

    manifest = {
        "version": args.version,
        "status": "LEARNING_GRAPH_V3",
        "question_status": REQUIRED_STATUS,
        "generated_at": generated_at,
        "revision": revision,
        "source_sha256": source_sha256,
        "source_sheet_id": args.source_sheet_id,
        "grammar_count": len(grammar),
        "vocabulary_count": len(vocabulary),
        "reading_count": len(reading_items),
        "total_item_count": len(all_questions),
        "reading_test_count": len(reading_tests),
        "concept_count": len(concepts),
        "edge_count": len(edges),
        "item_graph_count": len(item_graph),
        "adaptive_rule_count": len(adaptive_rules),
        "learner_state_count": len(learner_state),
    }
    write_json(args.output_dir / "manifest.json", manifest)
    print(json.dumps(manifest, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    try:
        raise SystemExit(main())
    except Exception as exc:  # noqa: BLE001
        print(f"ERROR: {exc}", file=sys.stderr)
        raise SystemExit(1)
